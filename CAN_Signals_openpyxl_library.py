import os
import sys
import time
import shutil
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.cell import get_column_letter
from openpyxl import load_workbook


directory = r'D:\VP2AutoTest\ROBOTFRAMEWORK\Automatic_tests_code\executable_testcases\can\nextgen\nextgen_variables'
directory_excel = r'D:\TABELE\\'
temp = directory_excel + "TEMP"
# directory = raw_input("Paste here path to the Carline variables directory (E.g. carlines\FGA_312\FGA_312_variables\n")
# directory_excel = raw_input("Paste here path to the directory, where do you want to create excel files")
# creating necessary folders
if not os.path.exists(temp):
   os.makedirs(temp)
count = 0

# Printing txt file names
for filename in os.listdir(directory):
    count += 1
    print filename

# Printing number of files in directory
if count > 0:
    print "\nThis folder contains %s files" % count
    time.sleep(1)
else:
    sys.exit("This directory is empty")

if count == 1:
    print("Creating %s Workbook" % count)
else:
    print("Creating %s Workbooks" % count)

# Creating xlsx files with three sheets
for filename in os.listdir(directory):
    # removing .txt from excel name
    new_name = filename.replace('.robot', "")
    wb = Workbook()
    ws = wb.active
    # adding 3 sheets to excel file
    ws.title = "Signals"
    ws2 = wb.create_sheet("Output")
    ws3 = wb.create_sheet("Values")
    # saving excel file
    wb.save(directory_excel + "%s.xlsx" % new_name)

# Erasing signal name and replace '}' to '\t' before uploading information to excel files
for filename in os.listdir(directory):
    with open(directory + '\\' + filename, 'r') as infile, open(temp + '\\' + filename, 'w') as outfile:
        # reading infile
        file = infile.read()
        # replacing unnecessary data in files e.g ${ipc_ack_DisplayInfoReceptionSts} to DisplayInfoReceptionSts
        replace = filename.lower().replace('.robot', "")
        outfile.write(file.replace('${' + replace + '_', "").replace('}', '\t').strip('*** Variables ***' + '\n'))
        infile.close()
        outfile.close()

# uploading data to excel file from txt file in temp folder
for filename in os.listdir(temp):
    # looking for this string in each line of txt file
    unwanted_char = "_"
    with open(temp + '\\' + filename) as txt_file:
        excel_file = (directory_excel + '\\' + filename.replace('.robot', '.xlsx'))
        wb = load_workbook(excel_file)
        # integers with proper row number for cells
        row_signals = 0
        row_values = 0
        # reading line by line txt file
        for line in txt_file:
            # looking for "_" in line, if there won't be "_" copy this line to first workbook sheet ("Signals")
            if unwanted_char not in line:
                row_signals += 1
                ws = wb.active
                ws2 = ws3 = wb.get_sheet_by_name("Output")
                signal, value = line.split('\t')
                ws.cell(row=row_signals, column=1).value = signal
                ws2.cell(row=row_signals, column=1).value = signal
                wb.save(excel_file)
            else:
                # if line contain "_" will be copy to last third sheet ("Values")
                row_values += 1
                ws3 = wb.get_sheet_by_name("Values")
                signal, value = line.split('\t')
                # first cell for signal attributes
                ws3.cell(row=row_values, column=1).value = signal
                # second cell for signal values
                ws3.cell(row=row_values, column=2).value = value
                wb.save(excel_file)

# removing TEMP file with txt file
shutil.rmtree(temp)

# creating formula for data validation
for filename in os.listdir(directory_excel):
    # variables
    count = 0
    split_string = "_"
    ws_row_values = ws.max_row
    ws3_row_values = ws3.max_row
    row_for_dv = 1
    # magic
    formula_for_dv = ""
    excel_file = (directory_excel + filename)
    wb = load_workbook(excel_file)
    ws = wb.active
    sheet_one_row = 0
    for ws_row_values in ws:
        sheet_one_row += 1
        sheet_two_row = 1
        sheet_one_signal_name = ws.cell(row=sheet_one_row, column=1).value
        print sheet_one_signal_name
        ws3 = wb.get_sheet_by_name("Values")
        for ws3_row_values in ws3:
            sheet_two_signal_name = ws3.cell(row=sheet_two_row, column=1).value
            stripped = sheet_two_signal_name.split('_')[0]
            if sheet_one_signal_name == stripped and count < 8:
                formula_for_dv += str(sheet_two_signal_name) + ','
                print ws3.cell(row=sheet_two_row, column=1)
                sheet_two_row += 1
                count += 1
            else:
                print "no"
                sheet_two_row += 1
        #print "to jest formula i zawiera %s " %formula_for_dv + '\n'
        dv = DataValidation(type="list", formula1='"%s"'%formula_for_dv, allow_blank=True)
        c1 = ws["B%s"%row_for_dv]
        dv.add(c1)
        ws.add_data_validation(dv)
        row_for_dv += 1
        formula_for_dv = ""
        count = 0
        wb.save(excel_file)










